import pandas as pd
import numpy as np
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
import streamlit as st

st.write("Gerador de sequenciamento da Pintura")

df = pd.read_csv(r"\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\PLANILHA DE CARGAS 2022 - Cargas.csv")

df1 = pd.read_excel(r"\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Base_carretas.xlsx")

#excluindo colunas inúteis

df = df.drop(columns=['PROGRAMAÇÃO DE CARGAS - Produção','Unnamed: 0','Unnamed: 5',
                       'Unnamed: 6','Unnamed: 4','Unnamed: 7','Unnamed: 23',
                       'Unnamed: 10','Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13',
                       'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16', 'Unnamed: 17',
                       'Unnamed: 18', 'Código:', 'PQ VE-001-000', 'Unnamed: 21',
                       'Unnamed: 22'])

#excluindo linhas inúteis

df = df.drop([0,1,2]).reset_index(drop=True)

#renomeando colunas

df.rename(columns={'Unnamed: 1': 'Datas', 'Unnamed: 2':'Carga',
                    'Unnamed: 8':'Recurso','Unnamed: 9':'Qtde'}, inplace = True)

df1.rename(columns={'    ': 'Recurso', 'PEÇA':'Peca'}, inplace = True)

count_linhas = df.shape[0]

#repetindo celula anterior na celula atual se for nulo

for i in range(1,count_linhas):
    if pd.isnull(df['Datas'][i]) == True:
        df['Datas'][i] = df['Datas'][i-1]

for i in range(1,count_linhas):
    if pd.isnull(df['Carga'][i]) == True:
        df['Carga'][i] = df['Carga'][i-1]

#apagando linha se existir algum valor nulo

df = df.dropna(axis=0)
df = df.reset_index(drop=True)

#df1 = df1.dropna(axis=0)

df1 = df1.dropna(subset=["Etapa2"])
df1 = df1.reset_index(drop=True)

df1 = df1.astype(str)

for data in range(0,df.shape[0]):
    df['Datas'][data] = df['Datas'][data][6:]

for d in range(0,df1.shape[0]):
        
    if len(df1['Código'][d]) == 5:
        df1['Código'][d] = '0' + df1['Código'][d]

#separando string por "-" e adicionando no dataframe antigo

tratando_coluna = df["Recurso"].str.split(" - ", n = 1, expand = True)

df['Recurso'] = tratando_coluna[0]

#tratando cores da string

df['Recurso_cor'] = df['Recurso']

df = df.reset_index(drop=True)

df_cores = pd.DataFrame({'Recurso_cor':['AN','VJ','LC','VM','AV','sem_cor'], 
                         'cor':['Azul','Verde','Laranja','Vermelho','Amarelo','Laranja']})

cores = ['AM','AN','VJ','LC','VM','AV']

for r in range(0,df.shape[0]):
    df['Recurso_cor'][r] = df['Recurso_cor'][r][len(df['Recurso_cor'][r])-3:len(df['Recurso_cor'][r])]
    df['Recurso_cor'] = df['Recurso_cor'].str.strip()
    
    if len(df['Recurso_cor'][r]) > 2:
        df['Recurso_cor'][r] = df['Recurso_cor'][r][1:3]
                
    if df['Recurso_cor'][r] not in cores:
        df['Recurso_cor'][r] = "LC"
     
df = pd.merge(df, df_cores, on=['Recurso_cor'], how='left')
   
#df['Recurso']=df['Recurso'].str.replace('AM','') 
df['Recurso']=df['Recurso'].str.replace('AN','') # Azul
df['Recurso']=df['Recurso'].str.replace('VJ','') # Verde
df['Recurso']=df['Recurso'].str.replace('LC','') # Laranja
df['Recurso']=df['Recurso'].str.replace('VM','') # Vermelho
df['Recurso']=df['Recurso'].str.replace('AV','') # Amarelo

#retirando espaços em branco no final da string

df['Recurso'] = df['Recurso'].str.strip()

#filtrando datas

for ano in range(0,len(df)):
    df['Ano'] = df['Datas'][ano][6:8]

df['Ano'] = df['Ano'].str.replace("22","2022")

for vinte_dois in range(0,len(df)):
    df['Datas'][vinte_dois] = df['Datas'][vinte_dois][0:6]
    df['Datas'][vinte_dois] = df['Datas'][vinte_dois] + df['Ano'][vinte_dois]

#tipo_filtro = input("Digite a data da carga: ")
tipo_filtro = st.sidebar.text_input("Digite a data no formato: dd/mm/aaaa")

datas_unique = pd.DataFrame(df['Datas'].unique())
    
escolha_data = (df['Datas'] == tipo_filtro)
filtro_data = df.loc[escolha_data]
filtro_data['Datas'] = pd.to_datetime(filtro_data.Datas)
   
#procv e trazendo as colunas que quero ver

tab_completa = pd.merge(filtro_data, df1, on=['Recurso'], how='left')

tab_completa['Código'] = tab_completa['Código'].astype(str)

tab_completa = tab_completa.reset_index(drop=True)

celulas_unique = pd.DataFrame(tab_completa['Célula'].unique())
celulas_unique = celulas_unique.dropna(axis=0)
celulas_unique.reset_index(drop=True)

recurso_unique =  pd.DataFrame(tab_completa['Recurso'].unique())
recurso_unique = recurso_unique.dropna(axis=0)

#tratando coluna de código

for t in range(0,tab_completa.shape[0]):
    
    if len(tab_completa['Código'][t]) == 5:
        tab_completa['Código'][t] = '0' + tab_completa['Código'][t][0:5]
        
    if len(tab_completa['Código'][t]) == 8:
        tab_completa['Código'][t] = tab_completa['Código'][t][0:6]
        
#criando coluna de quantidade total de itens

tab_completa['Qtde_x'] = tab_completa['Qtde_x'].str.replace(',','.')

tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(float)
tab_completa['Qtde_x'] = tab_completa['Qtde_x'].astype(int)

tab_completa = tab_completa.dropna(axis=0)

tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(float)
tab_completa['Qtde_y'] = tab_completa['Qtde_y'].astype(int)

tab_completa['Qtde_total'] = tab_completa['Qtde_x'] * tab_completa['Qtde_y']

tab_completa = tab_completa.drop(columns=['Carga','Recurso','Qtde_x','Qtde_y','Etapa','LEAD TIME','flag peça','Etapa2'])

tab_completa.columns
tab_completa = tab_completa.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum()
tab_completa.reset_index(inplace=True)

tab_completa.drop(tab_completa.loc[tab_completa['Célula']=='EIXO SIMPLES'].index, inplace=True)
tab_completa.reset_index(inplace=True, drop=True)

for t in range(0,len(tab_completa)):
    
    if tab_completa['Célula'][t] == 'FUEIRO' or \
    tab_completa['Célula'][t] == 'LATERAL' or \
    tab_completa['Célula'][t] == 'PLAT. TANQUE. CAÇAM.': 
        
        tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + tab_completa['Recurso_cor'][t]
    
    else:
        
        tab_completa['Recurso_cor'][t] = tab_completa['Código'][t] + 'CO' 
        tab_completa['cor'][t] = 'Cinza'
    
k = 9

cor_unique = tab_completa['cor'].unique()

for i in range(0,len(cor_unique)):
    
    wb = Workbook()
    wb = load_workbook(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\modelo_op_pintura.xlsx')
    ws = wb.active
    
    filtro_excel = (tab_completa['cor'] == cor_unique[i])
    filtrar = tab_completa.loc[filtro_excel]
    filtrar = filtrar.reset_index(drop=True)
    filtrar = filtrar.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum().reset_index()
    filtrar.sort_values(by=['Célula'], inplace=True)  
    filtrar = filtrar.reset_index(drop=True)
    
    if len(filtrar) > 21:
    
        for j in range(0,21):
         
            ws['F5'] = cor_unique[i] # nome da coluna é '0'
            ws['AD5'] = datetime.datetime.now() #  data de hoje
            ws['M4']  = tipo_filtro  # data da carga
            ws['B' + str(k)] = filtrar['Recurso_cor'][j]
            ws['G' + str(k)] = filtrar['Peca'][j]
            ws['AD' + str(k)] = filtrar['Qtde_total'][j]
            k = k + 1
            
            wb.template = False
            path = Path(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Sequenciamento Pintura\ ' + tipo_filtro.replace('/', '-')) 
            path.mkdir(parents=True, exist_ok=True)
            wb.save(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Sequenciamento Pintura\ ' + tipo_filtro.replace('/', '-') + '\ ' + cor_unique[i] +'.xlsx')    
        
        k = 9
        
        wb = Workbook()
        wb = load_workbook(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\modelo_op_pintura.xlsx')
        ws = wb.active
        
        filtro_excel = (tab_completa['cor'] == cor_unique[i])
        filtrar = tab_completa.loc[filtro_excel]
        filtrar = filtrar.reset_index(drop=True)
        filtrar = filtrar.groupby(['Código','Peca','Célula','Datas','Recurso_cor','cor']).sum().reset_index()
        filtrar.sort_values(by=['Célula'], inplace=True)  
        filtrar = filtrar.reset_index(drop=True)

        if len(filtrar) > 21:
            
            j = 21
            
            for j in range(21,len(filtrar)):
             
                ws['F5'] = cor_unique[i] # nome da coluna é '0'
                ws['AD5'] = datetime.datetime.now() #  data de hoje
                ws['M4']  = tipo_filtro  # data da carga
                ws['B' + str(k)] = filtrar['Recurso_cor'][j]
                ws['G' + str(k)] = filtrar['Peca'][j]
                ws['AD' + str(k)] = filtrar['Qtde_total'][j]
                k = k + 1
                
                wb.save(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Sequenciamento Pintura\ ' + tipo_filtro.replace('/', '-') + '\ ' + cor_unique[i] +'1.xlsx')
        
    else:
        
        j = 0
        k = 9
        for j in range(0,21-(21-len(filtrar))):
         
            ws['F5'] = cor_unique[i] # nome da coluna é '0'
            ws['AD5'] = datetime.datetime.now() #  data de hoje
            ws['M4']  = tipo_filtro  # data da carga
            ws['B' + str(k)] = filtrar['Recurso_cor'][j]
            ws['G' + str(k)] = filtrar['Peca'][j]
            ws['AD' + str(k)] = filtrar['Qtde_total'][j]
            k = k + 1
            
            wb.template = False
            path = Path(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Sequenciamento Pintura\ ' + tipo_filtro.replace('/', '-')) 
            path.mkdir(parents=True, exist_ok=True)
            wb.save(r'\\cemag-server\Cemag\Grupos\Usinagem\5.SEQUENCIAMENTO\Sequenciamento Pintura\ ' + tipo_filtro.replace('/', '-') + '\ ' + cor_unique[i] +'.xlsx')
            
        k = 9    